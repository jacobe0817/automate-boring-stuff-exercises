# python 3
# chore_scheduler.py

# takes the path to an .xlsx file where column A is email addresses and column B is chores
# if address column contains duplicates raise an exception
# if chore column contains ';' or the string 'NO CHORE', raise an exception
# if more chores than addresses, at least one address will have an alphabetical ; seperated string of chores
# if more addresses than chores, some people have no chore for the week
# if there is a schedule sheet, returns
# otherwise, adds a sheet for schedule (row 1 is email addresses, each subsequent row is the alphabetical ; separated string of chores assigned to each address for that week)
# forms a shuffled master list of dicts where each dict corresponds to a week, and dict[address] = assigned chores
# writes the master list to the schedule sheet

# can be run as a script or imported

import os
import openpyxl
import copy
import random
import itertools
from collections import defaultdict

def main():
    file_path = input('\nProvide a path to an Excel workbook with a sheet titled "inputs". The first column must be addresses, the second must be chores, and there should be no headers. Neither column may contain duplicates.\n\n> ')
    validate_file_path(file_path)
    generate_schedule(file_path)

def validate_file_path(file_path, allow_schedule_sheet=False):
    if not os.path.exists(file_path):
        raise Exception(f'Specified file path {file_path} does not exist.')
    
    if not file_path.endswith('.xlsx'):
        raise Exception('Specified file must end with ".xlsx".')
    
    wb = openpyxl.load_workbook(file_path)
    
    if not 'inputs' in wb.sheetnames:
        wb.close()
        raise Exception('Specified workbook has no sheet titled "inputs".')

    if not allow_schedule_sheet and 'schedule' in wb.sheetnames:
        wb.close()
        raise Exception('Specified workbook already contains a sheet titled "schedule". Delete or rename this sheet to continue.')
    
    ws = wb['inputs']
    addresses = [cell.value for cell in ws['A'] if cell.value is not None]
    chores = [cell.value for cell in ws['B'] if cell.value is not None]

    if len(addresses) == 0:
        wb.close()
        raise Exception('Address column must contain at least one value.')

    if len(chores) == 0:
        wb.close()
        raise Exception('Chore column must contain at least one value.')

    for chore in chores:
        if ';' in chore:
            wb.close()
            raise Exception('A chore cannot contain a semi-colon ";".')

    if 'NO CHORE' in chores:
        wb.close()
        raise Exception('Chore column cannot contain the value "NO CHORE".')

    if len(addresses) != len(set(addresses)):
        wb.close()
        raise Exception('Address column cannot contain duplicates.')

    if len(chores) != len(set(chores)):
        wb.close()
        raise Exception('Chore column cannot contain duplicates. If you want "laundry" listed twice, try "laundry A" and "laundry B".')

    wb.close()

def generate_schedule(file_path):
    wb = openpyxl.load_workbook(file_path)
    input_sheet = wb['inputs']
    addresses = [cell.value for cell in input_sheet['A'] if cell.value is not None].sort()
    chores = [cell.value for cell in input_sheet['B'] if cell.value is not None].sort()

    for i in range(len(chores) - len(addresses)):
        chores.append(f'NO CHORE {i + 1}')

    master_list = get_master_list(addresses, chores)

    wb.close()

# returns a list with length = len(addresses) where each element corresponds to the number of chores assigned to one address
def get_assignment_sizes(num_addresses, num_chores):
    address_chore_count = defaultdict(int)
    for i in range(num_chores):
        key = i % num_addresses
        address_chore_count[key] += 1
    return list(address_chore_count.values())

# use a recursive function to create all possible groupings
def all_possible_chore_groupings(chores, assignment_sizes, assignment_index=0):
    # print(f'called all_possible_chore_groupings with chores: {chores}, assignment_sizes: {assignment_sizes}, and index {assignment_index}')
    combinations = list(itertools.combinations(chores, assignment_sizes[assignment_index]))
    return_list = []
    if assignment_index < len(assignment_sizes) - 1:
        for combination in combinations:
            chores_copy = copy.copy(chores)
            for x in combination:
                chores_copy.remove(x)
            for remaining_list in all_possible_chore_groupings(chores_copy, assignment_sizes, assignment_index + 1):
                if isinstance(remaining_list, tuple):
                    remaining_list = [remaining_list]
                remaining_list = [combination] + remaining_list
                remaining_list.sort()
                if not remaining_list in return_list:
                    return_list.append(remaining_list)
        return return_list
    return combinations

# returns a shuffled master list of dicts where each dict corresponds to a week, and dict[address] = tuple of assigned chores
def get_master_list(addresses, chores):
    addresses = copy.copy(addresses)
    chores = copy.copy(chores)

    assignment_sizes = get_assignment_sizes(len(addresses), len(chores))
    possible_chore_groupings = all_possible_chore_groupings(chores, assignment_sizes)

    print(len(possible_chore_groupings))
    master_list = []
    for grouping in possible_chore_groupings:
        offset = 0
        for i in range(len(addresses)):
            week_dict = {}
            week_dict[addresses[i + offset % len(addresses)]] = 
            
    print(len(master_list))
    random.shuffle(master_list)

    return master_list

if __name__ == '__main__':
    get_master_list(['a', 'b', 'c', 'd'], [1,2,3,4,5,6,7,8,9,10])