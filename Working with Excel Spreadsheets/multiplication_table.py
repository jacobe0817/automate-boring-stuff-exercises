import sys
import openpyxl
import os

folder_name = 'multiplication tables'

# process command line argument
try:
    num = sys.argv[1]
    num = int(num)
except:
    print('\nMust provide an integer argument')
    sys.exit()

str_num = str(num)
file_name = str_num + ' x ' + str_num
file_path = os.path.join(folder_name, file_name + '.xlsx')

# make directory if it doen't exist already
if not os.path.exists(folder_name):
    os.mkdir(folder_name)
    print(f'\nDirectory "{os.path.abspath(folder_name)}" created')

# initialize workbook
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.freeze_panes = 'B2'

# initialize bold_font
bold_font = openpyxl.styles.Font(name='Calibri', bold=True)
sheet.cell(row=1, column=1, value=file_name).font = bold_font

# loop to populate table
for x in range(1, num + 1):
    row_cell = sheet.cell(row=1, column=x + 1, value=x)
    col_cell = sheet.cell(row=x + 1, column=1, value=x)
    
    row_cell.font = bold_font
    col_cell.font = bold_font

for x in range(1, num + 1):
    for y in range(1, num + 1):
        _ = sheet.cell(row=x + 1, column=y + 1, value=x * y)
    
# save workbook
workbook.save(file_path)
print(f'\nMultiplication table saved to {file_path}')